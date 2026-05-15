from fastapi import FastAPI, UploadFile, File, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
from sqlalchemy import create_engine, Column, String, DateTime, Text, Float, JSON
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
import httpx, os, uuid, json, io, re
from pathlib import Path
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import PyPDF2
from pdf2image import convert_from_bytes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

app = FastAPI(title="FacturaAI API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./facturas.db")
GROQ_KEY     = os.getenv("GROQ_API_KEY", "")
UPLOAD_DIR   = Path("/app/uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+psycopg2://") if DATABASE_URL.startswith("postgresql://") else DATABASE_URL
engine       = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
Base         = declarative_base()

# ── Models ───────────────────────────────────────────────
class Factura(Base):
    __tablename__ = "facturas"
    id                    = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    filename              = Column(String)
    tipo                  = Column(String)
    proveedor_nombre      = Column(String, nullable=True)
    proveedor_cedula      = Column(String, nullable=True)
    proveedor_telefono    = Column(String, nullable=True)
    numero_factura        = Column(String, nullable=True)
    fecha_factura         = Column(String, nullable=True)
    monto_total           = Column(Float,  nullable=True)
    sociedad_detectada    = Column(String, nullable=True)
    cedula_detectada      = Column(String, nullable=True)
    estado_validacion     = Column(String, default="pendiente")
    liquidador_nombre     = Column(String, nullable=True)
    liquidador_codigo     = Column(String, nullable=True)
    ocr_confidence        = Column(Float,  nullable=True)
    texto_extraido        = Column(Text,   nullable=True)
    datos_raw             = Column(JSON,   nullable=True)
    created_at            = Column(DateTime, default=datetime.utcnow)

class EmpresaConfig(Base):
    __tablename__ = "empresa_config"
    id             = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    nombre_oficial = Column(String)
    variaciones    = Column(JSON, default=[])
    cedula_juridica = Column(String)
    updated_at     = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(bind=engine)

def init_config():
    db = SessionLocal()
    if not db.query(EmpresaConfig).first():
        config = EmpresaConfig(
            nombre_oficial="Mi Empresa S.A.",
            variaciones=["Mi Empresa", "EMPRESA SA", "Mi Empresa Sociedad Anónima"],
            cedula_juridica="3-101-000000"
        )
        db.add(config); db.commit()
    db.close()

init_config()

def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

def clean_number(value: any) -> float:
    if value is None: return None
    if isinstance(value, (int, float)): return float(value)
    s = str(value).strip()
    if s.count(',') == 1 and '.' not in s:
        s = s.replace(',', '.')
    s = s.replace('.', '')
    s = re.sub(r'[^\d\.\-]', '', s)
    try:
        return float(s)
    except:
        return None

def preprocess_image(img: Image.Image) -> Image.Image:
    img = img.convert('L')
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    return img

def extract_text_from_image(img: Image.Image) -> tuple[str, float]:
    processed = preprocess_image(img)
    data = pytesseract.image_to_data(processed, lang='spa+eng', output_type=pytesseract.Output.DICT)
    text = pytesseract.image_to_string(processed, lang='spa+eng')
    confidences = [int(c) for c in data['conf'] if int(c) > 0]
    avg_conf = sum(confidences) / len(confidences) if confidences else 0
    return text, avg_conf

def extract_text_from_pdf(pdf_bytes: bytes) -> tuple[str, float]:
    text = ""
    try:
        reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            t = page.extract_text()
            if t: text += t + "\n"
    except: pass
    if len(text.strip()) < 50:
        try:
            images = convert_from_bytes(pdf_bytes, dpi=200)
            for img in images:
                t, _ = extract_text_from_image(img)
                text += t + "\n"
        except: pass
    return text, 85.0

async def analyze_with_groq(text: str, empresa_config: dict) -> dict:
    prompt = f"""Analiza el siguiente texto extraído de una factura y extrae la información en formato JSON.

Texto de la factura:
{text[:3000]}

Configuración de la empresa receptora:
- Nombre oficial: {empresa_config['nombre_oficial']}
- Variaciones aceptadas: {empresa_config['variaciones']}
- Cédula jurídica: {empresa_config['cedula_juridica']}

Extrae y devuelve SOLO un JSON válido con esta estructura exacta:
{{
  "proveedor_nombre": "nombre del proveedor/emisor de la factura",
  "proveedor_cedula": "cédula jurídica o física del proveedor",
  "proveedor_telefono": "teléfono si existe, null si no",
  "numero_factura": "número o consecutivo de factura",
  "fecha_factura": "fecha en formato DD/MM/YYYY",
  "monto_total": 0.00,
  "sociedad_detectada": "nombre de la empresa receptora encontrado en la factura",
  "cedula_detectada": "cédula jurídica de la empresa receptora encontrada",
  "estado_validacion": "valida|invalida|revisar",
  "razon_estado": "explicación breve del estado"
}}

Para el estado_validacion:
- "valida": el nombre y cédula de la empresa coinciden con la configuración
- "invalida": la factura claramente NO pertenece a la empresa
- "revisar": hay datos pero no son completamente claros

Responde SOLO con el JSON, sin texto adicional. Asegúrate de que el campo monto_total sea un número sin puntos de miles (usa punto decimal si hay céntimos)."""

    async with httpx.AsyncClient(timeout=30) as client:
        res = await client.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {GROQ_KEY}", "Content-Type": "application/json"},
            json={"model": "llama-3.3-70b-versatile", "messages": [{"role":"user","content":prompt}], "temperature":0.1}
        )
        data = res.json()
        if "choices" not in data: raise Exception(str(data))
        content = data["choices"][0]["message"]["content"]
        clean = content.replace("```json","").replace("```","").strip()
        result = json.loads(clean)
        if 'monto_total' in result:
            result['monto_total'] = clean_number(result['monto_total'])
        return result

class EmpresaConfigBody(BaseModel):
    nombre_oficial: str
    variaciones: List[str]
    cedula_juridica: str

class LiquidadorBody(BaseModel):
    factura_id: str
    liquidador_nombre: str
    liquidador_codigo: str

# ── Routes ────────────────────────────────────────────────
@app.get("/")
def root(): return {"status": "FacturaAI running", "version": "1.0.0"}

@app.get("/config")
def get_config(db: Session = Depends(get_db)):
    config = db.query(EmpresaConfig).first()
    if not config: raise HTTPException(404, "No config found")
    return {"id":config.id,"nombre_oficial":config.nombre_oficial,"variaciones":config.variaciones,"cedula_juridica":config.cedula_juridica}

@app.put("/config")
def update_config(body: EmpresaConfigBody, db: Session = Depends(get_db)):
    config = db.query(EmpresaConfig).first()
    if not config:
        config = EmpresaConfig()
        db.add(config)
    config.nombre_oficial = body.nombre_oficial
    config.variaciones = body.variaciones
    config.cedula_juridica = body.cedula_juridica
    config.updated_at = datetime.utcnow()
    db.commit(); db.refresh(config)
    return {"ok": True, "config": {"nombre_oficial":config.nombre_oficial,"variaciones":config.variaciones,"cedula_juridica":config.cedula_juridica}}

@app.post("/facturas/upload")
async def upload_factura(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    ext = Path(file.filename).suffix.lower()
    file_id = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{file_id}{ext}"
    with open(file_path, "wb") as f: f.write(content)

    config = db.query(EmpresaConfig).first()
    empresa_data = {"nombre_oficial":config.nombre_oficial,"variaciones":config.variaciones,"cedula_juridica":config.cedula_juridica}

    if ext in ['.jpg','.jpeg','.png','.webp']:
        img = Image.open(io.BytesIO(content))
        text, conf = extract_text_from_image(img)
        tipo = "fisica"
    elif ext == '.pdf':
        text, conf = extract_text_from_pdf(content)
        tipo = "electronica"
    else:
        raise HTTPException(400, "Formato no soportado. Use JPG, PNG o PDF.")

    try:
        extracted = await analyze_with_groq(text, empresa_data)
    except Exception as e:
        extracted = {"proveedor_nombre":None,"proveedor_cedula":None,"proveedor_telefono":None,"numero_factura":None,"fecha_factura":None,"monto_total":None,"sociedad_detectada":None,"cedula_detectada":None,"estado_validacion":"revisar","razon_estado":f"Error AI: {str(e)}"}

    factura = Factura(
        filename=file.filename,
        tipo=tipo,
        proveedor_nombre=extracted.get("proveedor_nombre"),
        proveedor_cedula=extracted.get("proveedor_cedula"),
        proveedor_telefono=extracted.get("proveedor_telefono"),
        numero_factura=extracted.get("numero_factura"),
        fecha_factura=extracted.get("fecha_factura"),
        monto_total=extracted.get("monto_total"),
        sociedad_detectada=extracted.get("sociedad_detectada"),
        cedula_detectada=extracted.get("cedula_detectada"),
        estado_validacion=extracted.get("estado_validacion","revisar"),
        ocr_confidence=round(conf,1),
        texto_extraido=text[:2000],
        datos_raw=extracted,
    )
    db.add(factura); db.commit(); db.refresh(factura)
    return factura

@app.get("/facturas")
def list_facturas(estado: Optional[str] = None, proveedor: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(Factura)
    if estado: q = q.filter(Factura.estado_validacion == estado)
    if proveedor: q = q.filter(Factura.proveedor_nombre.ilike(f"%{proveedor}%"))
    return q.order_by(Factura.created_at.desc()).all()

@app.get("/facturas/{factura_id}")
def get_factura(factura_id: str, db: Session = Depends(get_db)):
    f = db.query(Factura).filter(Factura.id == factura_id).first()
    if not f: raise HTTPException(404, "Factura no encontrada")
    return f

@app.put("/facturas/{factura_id}/liquidar")
def liquidar_factura(factura_id: str, body: LiquidadorBody, db: Session = Depends(get_db)):
    f = db.query(Factura).filter(Factura.id == factura_id).first()
    if not f: raise HTTPException(404, "Factura no encontrada")
    f.liquidador_nombre = body.liquidador_nombre
    f.liquidador_codigo = body.liquidador_codigo
    db.commit(); db.refresh(f)
    return f

@app.put("/facturas/{factura_id}/estado")
def update_estado(factura_id: str, estado: str, db: Session = Depends(get_db)):
    f = db.query(Factura).filter(Factura.id == factura_id).first()
    if not f: raise HTTPException(404, "Factura no encontrada")
    if estado not in ["valida","invalida","revisar"]: raise HTTPException(400, "Estado inválido")
    f.estado_validacion = estado
    db.commit(); db.refresh(f)
    return f

@app.delete("/facturas/{factura_id}")
def delete_factura(factura_id: str, db: Session = Depends(get_db)):
    f = db.query(Factura).filter(Factura.id == factura_id).first()
    if not f: raise HTTPException(404, "Factura no encontrada")
    db.delete(f); db.commit()
    return {"deleted": True}

@app.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    total = db.query(Factura).count()
    validas = db.query(Factura).filter(Factura.estado_validacion=="valida").count()
    invalidas = db.query(Factura).filter(Factura.estado_validacion=="invalida").count()
    revisar = db.query(Factura).filter(Factura.estado_validacion=="revisar").count()
    monto_t = db.query(Factura).filter(Factura.monto_total != None).all()
    total_m = sum(f.monto_total for f in monto_t if f.monto_total)
    return {"total":total,"validas":validas,"invalidas":invalidas,"revisar":revisar,"total_monto":round(total_m,2),"pct_invalidas":round(invalidas/total*100,1) if total else 0}

@app.get("/export/excel")
def export_excel(estado: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(Factura)
    if estado: q = q.filter(Factura.estado_validacion == estado)
    facturas = q.order_by(Factura.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = ["Proveedor","Cédula Proveedor","Teléfono","Nº Factura","Fecha","Monto Total","Tipo","Sociedad Detectada","Cédula Detectada","Estado","OCR %","Liquidador","Código Empleado","Fecha Proceso"]
    header_fill = PatternFill(start_color="1e1b4b", end_color="1e1b4b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    estado_colors = {"valida":"c6efce","invalida":"ffc7ce","revisar":"ffeb9c"}

    for row, f in enumerate(facturas, 2):
        values = [f.proveedor_nombre,f.proveedor_cedula,f.proveedor_telefono,f.numero_factura,f.fecha_factura,f.monto_total,f.tipo,f.sociedad_detectada,f.cedula_detectada,f.estado_validacion,f.ocr_confidence,f.liquidador_nombre,f.liquidador_codigo,f.created_at.strftime("%d/%m/%Y %H:%M") if f.created_at else ""]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            if col == 10 and f.estado_validacion in estado_colors:
                cell.fill = PatternFill(start_color=estado_colors[f.estado_validacion], end_color=estado_colors[f.estado_validacion], fill_type="solid")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 18

    path = UPLOAD_DIR / "export.xlsx"
    wb.save(path)
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="facturas_export.xlsx")

# ── Exportación a SAP (formato exacto con colores, bordes, celdas editables) ──
def create_sap_excel(factura: Factura) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # ── Colores exactos de la imagen ─────────────────────
    PURPLE_DARK  = "7030A0"   # encabezados fila 4 y fila 9
    PURPLE_LIGHT = "D9B3FF"   # filas 11, 12 (Conceptos/Cuenta/Ceco vacíos)
    BLUE_LIGHT   = "ADD8E6"   # celdas editables azul claro
    YELLOW       = "FFFF00"   # celda Acreedor valor (B3)
    PINK_TITLE   = "FFE4E1"   # fondo título D1:F2
    GRAY_HEADER  = "D9D9D9"   # encabezados fila 9

    # ── Anchos de columna (A-K) ──────────────────────────
    col_widths = [12, 16, 14, 30, 14, 14, 14, 8, 14, 8, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Altura de filas ──────────────────────────────────
    for r in range(1, 16):
        ws.row_dimensions[r].height = 18

    # ── Fila 1-2: Título ─────────────────────────────────
    ws.merge_cells('D1:F1')
    ws['D1'] = "FORMATO DE LIQUIDACIÓN"
    ws['D1'].font = Font(bold=True, size=12)
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].fill = PatternFill(start_color=PINK_TITLE, end_color=PINK_TITLE, fill_type="solid")

    ws.merge_cells('D2:F2')
    ws['D2'] = "ENCUENTROS SE"
    ws['D2'].font = Font(bold=True, size=11)
    ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D2'].fill = PatternFill(start_color=PINK_TITLE, end_color=PINK_TITLE, fill_type="solid")

    # ── Fila 3: Acreedor ─────────────────────────────────
    ws['A3'] = "Acreedor"
    ws['A3'].font = Font(bold=True)

    ws['B3'] = factura.proveedor_cedula or "400022039"
    ws['B3'].fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    ws['B3'].font = Font(bold=True)
    ws['B3'].alignment = Alignment(horizontal='center')

    # ── Fila 4: Encabezados SAP (fila morada) ────────────
    sap_headers = {
        'A4': 'Cta mayor',
        'B4': 'D/H',
        'C4': 'Importe moneda',
        'D4': 'Asignación',
        'E4': 'In',
        'F4': 'Texto',
        'G4': 'Pe',
        'H4': 'Ej',
        'I4': 'M.',
        'J4': 'Ne..',
        'K4': 'Centro de coste',
    }
    for cell_ref, value in sap_headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.fill = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Fila 5: Datos SAP ────────────────────────────────
    importe    = factura.monto_total if factura.monto_total else 0
    asignacion = f"ERC-{factura.id[:8]}" if factura.id else "ERC-2039-26-007"
    texto      = f"ENCUENTRO SE {(factura.filename or '')[:15].upper()}"
    anio       = str(datetime.utcnow().year)
    cta_mayor  = "610303001"
    ceco       = "10030601"

    ws['A5'] = cta_mayor
    ws['A5'].font = Font(size=10)

    ws['C5'] = importe
    ws['C5'].number_format = '#,##0'
    ws['C5'].font = Font(size=10)

    ws['D5'] = asignacion
    ws['D5'].font = Font(size=10)

    ws['E5'] = "I6"
    ws['E5'].font = Font(size=10)

    ws['F5'] = texto
    ws['F5'].font = Font(size=10)

    ws['H5'] = anio
    ws['H5'].font = Font(size=10)

    ws['K5'] = ceco
    ws['K5'].font = Font(size=10)

    # ── Filas 6-8: vacías ────────────────────────────────
    # (dejar en blanco)

    # ── Fila 9: Encabezados tabla (morado) ───────────────
    table_headers = {
        'A9': 'Fecha',
        'B9': 'Conceptos',
        'C9': '# Documento',
        'D9': 'Nombre del proveedor',
        'E9': 'Cuenta',
        'F9': '',
        'G9': 'Ceco',
        'H9': 'OE',
        'I9': 'MONTO',
        'J9': 'IVA',
        'K9': 'N° SAP',
    }
    for cell_ref, value in table_headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.fill = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Fila 10: Datos principales (azul claro editable) ─
    fecha      = factura.fecha_factura or datetime.utcnow().strftime("%d.%m.%Y")
    concepto   = "ALIMENTACIÓN"
    nro_doc    = factura.numero_factura or ""
    proveedor  = factura.proveedor_nombre or ""
    sap_nro    = "20005235"

    row10_data = {
        'A10': fecha,
        'B10': concepto,
        'C10': nro_doc,
        'D10': proveedor,
        'E10': cta_mayor,
        'F10': '',
        'G10': ceco,
        'H10': '',
        'I10': importe,
        'J10': 0,
        'K10': sap_nro,
    }
    for cell_ref, value in row10_data.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical='center')

    # Celda K10 — N° SAP en rojo/negrita como en la imagen
    ws['K10'].font = Font(bold=True, color="FF0000", size=13)
    ws['K10'].alignment = Alignment(horizontal='right', vertical='center')

    # Monto I10 con símbolo ₡
    ws['I10'].number_format = '₡#,##0'
    ws['I10'].alignment = Alignment(horizontal='right')

    # ── Filas 11-12: solo Conceptos, Cuenta y Ceco (morado claro) ──
    for row in [11, 12]:
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws[f'{col_letter}{row}'].fill = PatternFill(start_color="EAD7FF", end_color="EAD7FF", fill_type="solid")

        ws[f'B{row}'] = concepto
        ws[f'B{row}'].font = Font(size=10)
        ws[f'B{row}'].fill = PatternFill(start_color="EAD7FF", end_color="EAD7FF", fill_type="solid")

        ws[f'E{row}'] = cta_mayor
        ws[f'E{row}'].font = Font(size=10)
        ws[f'E{row}'].fill = PatternFill(start_color="EAD7FF", end_color="EAD7FF", fill_type="solid")

        ws[f'G{row}'] = ceco
        ws[f'G{row}'].font = Font(size=10)
        ws[f'G{row}'].fill = PatternFill(start_color="EAD7FF", end_color="EAD7FF", fill_type="solid")

        ws[f'I{row}'].number_format = '₡#,##0'
        ws[f'I{row}'].alignment = Alignment(horizontal='right')

    # ── Fila 13: Total ───────────────────────────────────
    ws['G13'] = "Total"
    ws['G13'].font = Font(bold=False, size=10)
    ws['G13'].alignment = Alignment(horizontal='right')

    ws['I13'] = f"=I10+I11+I12"
    ws['I13'].number_format = '₡#,##0'
    ws['I13'].alignment = Alignment(horizontal='right')
    ws['I13'].font = Font(size=10)

    # ── Fila 14: Deposito ────────────────────────────────
    ws['H14'] = "Deposito"
    ws['H14'].font = Font(size=10)
    ws['H14'].alignment = Alignment(horizontal='right')

    ws['I14'] = importe
    ws['I14'].number_format = '₡#,##0'
    ws['I14'].alignment = Alignment(horizontal='right')
    ws['I14'].font = Font(size=10)

    # ── Fila 15: Diferencia ──────────────────────────────
    ws['H15'] = "Diferencia"
    ws['H15'].font = Font(size=10)
    ws['H15'].alignment = Alignment(horizontal='right')

    ws['I15'] = "=I13-I14"
    ws['I15'].number_format = '₡#,##0.00'
    ws['I15'].alignment = Alignment(horizontal='right')
    ws['I15'].font = Font(size=10)

    # ── Bordes en tabla (filas 9-15, cols A-K) ───────────
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(9, 16):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = border

    # ── Bordes en sección SAP (filas 4-5, cols A-K) ──────
    for row in range(4, 6):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = border

    # ── Celdas editables azul claro ──────────────────────
    editable = ['A10','B10','C10','D10','I10','I11','I12','B11','B12']
    for ref in editable:
        ws[ref].fill = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
        ws[ref].protection = Protection(locked=False)

    # ── Proteger hoja ────────────────────────────────────
    ws.protection.sheet = True
    ws.protection.password = ""

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.get("/export/sap/{factura_id}")
def export_to_sap(factura_id: str, db: Session = Depends(get_db)):
    factura = db.query(Factura).filter(Factura.id == factura_id).first()
    if not factura:
        raise HTTPException(404, "Factura no encontrada")
    excel_file = create_sap_excel(factura)
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=factura_{factura_id}_sap.xlsx"}
    )
